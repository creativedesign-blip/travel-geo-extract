"""Import 大都會旅遊 Excel master file into the 5 CSV knowledge base.

The Excel sheet `國家Mapping表` is the *authoritative* maintenance surface
(per-country row with status/review fields). This script reads it and
incrementally updates:

  - locations.csv      : add new country roots if missing
  - aliases.csv        : add new alias -> country mappings for strong keywords
                         that look like place / island / region names
  - landmark_index.csv : add strong keywords that look like landmarks
                         (temple, palace, park, brand, building...)
  - disambig_rules.csv : add SKIP rules for each row's exclusion keywords

Sheet `弱信號排除詞庫` is dumped into `weak_signals.csv` verbatim.

Usage:
    python import_xlsx.py [--xlsx <path>] [--dry-run]

Default --xlsx path: C:\\Users\\User\\Desktop\\2d3d\\旅遊表\\旅遊DM國家關鍵字排除Mapping表.xlsx
"""

from __future__ import annotations

import argparse
import csv
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))

from load_kb import load_kb  # noqa: E402

try:
    from openpyxl import load_workbook  # type: ignore
except ImportError as e:
    raise SystemExit(
        "openpyxl is required. pip install openpyxl"
    ) from e


DATA_DIR = Path(__file__).resolve().parent.parent / "data"

DEFAULT_XLSX = Path(r"C:\Users\User\Desktop\2d3d\旅遊表\旅遊DM國家關鍵字排除Mapping表.xlsx")

# P2-10: enforced schema for write protection. If a CSV header drifts (e.g. business
# manually reordered columns in Excel), refuse to write rather than silently misaligning.
EXPECTED_FIELDS = {
    "locations.csv": ["id", "name", "en_name", "ja_name", "type", "parent_id", "sort_order", "active", "notes"],
    "aliases.csv": ["alias", "location_id", "alias_type", "confidence", "notes"],
    "landmark_index.csv": ["landmark", "alias", "location_id", "type", "country_id", "notes"],
    "disambig_rules.csv": ["ambiguous_term", "context_regex", "window", "resolution", "priority", "reason"],
    "role_keywords.csv": ["keyword", "role", "window", "direction", "weight", "requires_context", "notes"],
    "weak_signals.csv": ["weak_signal", "naive_country", "possible_countries", "required_strong", "guidance", "min_context", "source"],
}

# 國家名 (Excel 寫法) -> ISO 3166-1 alpha-2 (skill 內部 id)
COUNTRY_ID = {
    "日本": "JP", "韓國": "KR",
    "越南": "VN", "泰國": "TH", "新加坡": "SG", "馬來西亞": "MY",
    "印尼": "ID", "菲律賓": "PH", "台灣": "TW", "中國": "CN",
    "香港": "HK", "澳門": "MO", "柬埔寨": "KH",
    "土耳其": "TR", "阿聯酋/杜拜": "AE", "阿聯酋": "AE",
    "義大利": "IT", "法國": "FR", "瑞士": "CH", "奧地利": "AT",
    "德國": "DE", "英國": "GB", "荷蘭": "NL", "西班牙": "ES",
    "葡萄牙": "PT", "希臘": "GR", "捷克": "CZ", "匈牙利": "HU",
    "克羅埃西亞": "HR",
    "埃及": "EG", "摩洛哥": "MA", "南非": "ZA",
    "澳洲": "AU", "紐西蘭": "NZ",
    "美國": "US", "加拿大": "CA", "墨西哥": "MX",
    "冰島": "IS", "芬蘭": "FI", "挪威": "NO",
    "印度": "IN", "尼泊爾": "NP",
}

# English names for new countries (used in locations.csv en_name column)
COUNTRY_EN = {
    "SG": "Singapore", "MY": "Malaysia", "HK": "Hong Kong", "MO": "Macau",
    "KH": "Cambodia", "TR": "Turkey", "AE": "UAE",
    "IT": "Italy", "FR": "France", "CH": "Switzerland", "AT": "Austria",
    "DE": "Germany", "GB": "United Kingdom", "NL": "Netherlands",
    "ES": "Spain", "PT": "Portugal", "GR": "Greece", "CZ": "Czech Republic",
    "HU": "Hungary", "HR": "Croatia",
    "EG": "Egypt", "MA": "Morocco", "ZA": "South Africa",
    "AU": "Australia", "NZ": "New Zealand",
    "US": "United States", "CA": "Canada", "MX": "Mexico",
    "IS": "Iceland", "FI": "Finland", "NO": "Norway",
    "IN": "India", "NP": "Nepal",
}

# Heuristic: terms ending in these suffixes are treated as landmarks (not city aliases)
LANDMARK_SUFFIXES = (
    "寺", "宮", "塔", "塢", "館", "樓", "橋", "廟", "神社", "城", "堡",
    "園", "公園", "山", "島", "灘", "瀑布", "湖", "河", "海岸",
    "Tower", "Palace", "Bridge", "Park", "Garden", "Museum",
    "Mall", "Centre", "Center", "Cathedral", "Square", "Plaza",
    "World", "Land", "World Wonders", "VinWonders",
)
LANDMARK_KEYWORDS = (
    "迪士尼", "環球影城", "USJ", "USS", "Universal", "Disney",
    "樂園", "水族館", "博物館", "教堂", "皇宮", "清真寺",
    "Grand World", "Vinpearl", "DubaiMall", "Dubai Mall",
    "金沙", "棕櫚島", "帆船", "魚尾獅", "雙子星", "101", "麥當勞",
)


def split_keywords(cell: str | None) -> list[str]:
    if not cell:
        return []
    # Excel uses 「、」 as separator; also accept "," and "/" and "；"
    parts = []
    buf = ""
    for ch in cell:
        if ch in "、,;；\n":
            if buf.strip():
                parts.append(buf.strip())
            buf = ""
        else:
            buf += ch
    if buf.strip():
        parts.append(buf.strip())
    return parts


def looks_like_landmark(term: str) -> bool:
    for suf in LANDMARK_SUFFIXES:
        if term.endswith(suf) and len(term) > len(suf):
            return True
    return any(kw in term for kw in LANDMARK_KEYWORDS)


def read_rows(path: Path) -> tuple[list[str], list[dict]]:
    with path.open(encoding="utf-8-sig", newline="") as fh:
        reader = csv.DictReader(fh)
        rows = list(reader)
        fields = reader.fieldnames or []
    return fields, rows


def _assert_header(path: Path, fieldnames: list[str]) -> None:
    """P2-10: Refuse to write if header doesn't match the expected schema."""
    expected = EXPECTED_FIELDS.get(path.name)
    if expected is None:
        return  # no schema declared for this file
    if list(fieldnames) != expected:
        raise SystemExit(
            f"\nCSV header drift in {path.name}:\n"
            f"  expected: {expected}\n"
            f"  actual:   {fieldnames}\n"
            f"Refusing to write to avoid column-misaligned rows. "
            f"Fix the header in the CSV before re-running import."
        )


def append_csv(path: Path, fieldnames: list[str], rows: list[dict]) -> None:
    if not rows:
        return
    _assert_header(path, fieldnames)
    with path.open("a", encoding="utf-8", newline="") as fh:
        writer = csv.DictWriter(fh, fieldnames=fieldnames, quoting=csv.QUOTE_MINIMAL)
        writer.writerows(rows)


def _load_curated_weak_set(xlsx_path: Path) -> set[str]:
    """Pre-load weak signal Sheet 2 so we know which strong-keyword tokens to skip.

    Reason: 「迪士尼」「極光」「魔戒」 are cross-country marketing terms; if Excel's
    per-country rows also list them as strong keywords, we DON'T want to pin them
    as alias/landmark for any single country.
    """
    wb = load_workbook(str(xlsx_path), data_only=True)
    if "弱信號排除詞庫" not in wb.sheetnames:
        return set()
    ws = wb["弱信號排除詞庫"]
    skip_set: set[str] = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        for part in split_keywords(row[0]):
            if part:
                skip_set.add(part)
    return skip_set


def import_country_sheet(xlsx_path: Path, kb, dry_run: bool) -> dict:
    wb = load_workbook(str(xlsx_path), data_only=True)
    ws = wb["國家Mapping表"]

    # collect what already exists to avoid duplicates
    existing_locations = {loc.id for loc in kb.locations.values()}
    existing_alias_pairs = {(a.alias, a.location_id) for a in kb.aliases}
    existing_landmarks = set(kb.landmark_index.keys())
    weak_skip = _load_curated_weak_set(xlsx_path)

    new_locations: list[dict] = []
    new_aliases: list[dict] = []
    new_landmarks: list[dict] = []

    # weak signal aggregation across all countries (key = weak_signal text)
    # value: {naive_country: str, possible_countries: set, required_strong: list, guidance: str}
    weak_agg: dict[str, dict] = {}

    # Excel sheet has 3 header rows; data starts at row 5 (1-indexed)
    for row in ws.iter_rows(min_row=5, values_only=True):
        country_name = (row[0] or "").strip() if row[0] else ""
        region_group = (row[1] or "").strip() if row[1] else ""
        strong_kw = row[2] or ""
        weak_kw = row[3] or ""
        # row[4]  排除邏輯 / 判斷說明
        # row[5]  建議 AI 判斷規則
        # row[6]  預設信心度
        # row[7]  需人工覆核條件
        status = (row[8] or "啟用").strip() if row[8] else "啟用"

        if not country_name or country_name == "國家/地區":
            continue
        if status == "停用":
            continue

        cid = COUNTRY_ID.get(country_name)
        if not cid:
            print(f"  skip unknown country: {country_name}", file=sys.stderr)
            continue

        # 1. locations.csv — add country root if missing
        if cid not in existing_locations:
            new_locations.append({
                "id": cid,
                "name": country_name,
                "en_name": COUNTRY_EN.get(cid, ""),
                "ja_name": "",
                "type": "country",
                "parent_id": "",
                "sort_order": "99",
                "active": "1",
                "notes": f"從 Excel 匯入 ({region_group})",
            })
            existing_locations.add(cid)
            # also register a Chinese alias auto-mapping in alias_index (via canonical name)

        # 2. aliases.csv / landmark_index.csv — classify each strong keyword
        for kw in split_keywords(strong_kw):
            # skip the country name itself; load_kb already auto-registers it
            if kw == country_name:
                continue
            # skip cross-country marketing terms — they belong in weak_signals only
            # (otherwise 「迪士尼」 from US row gets pinned as US-only landmark)
            if kw in weak_skip:
                continue
            # ignore if already mapped to this country (by canonical name, alias, or landmark)
            if (kw, cid) in existing_alias_pairs:
                continue
            if kw in kb.alias_index:
                # already maps to *something* — leave as is, avoid double mapping
                continue
            if kw in existing_landmarks:
                continue
            if looks_like_landmark(kw):
                # add to landmark_index, pinned to country root for now
                # (future enhancement: detect city and pin there)
                new_landmarks.append({
                    "landmark": kw,
                    "alias": "",
                    "location_id": cid,
                    "type": "landmark",
                    "country_id": cid,
                    "notes": "從 Excel 匯入 (掛國家根，待細化到城市)",
                })
                existing_landmarks.add(kw)
            else:
                # treat as alias for the country
                new_aliases.append({
                    "alias": kw,
                    "location_id": cid,
                    "alias_type": "landmark" if "島" in kw or "山" in kw else "nickname",
                    "confidence": "0.85",
                    "notes": "從 Excel 匯入",
                })
                existing_alias_pairs.add((kw, cid))

        # 3. weak_signals aggregation — accumulate possible countries per signal.
        # Exclusion keywords are NOT written to disambig_rules.csv (which would
        # unconditionally SKIP them); they go to weak_signals.csv so extract.py
        # only flags them when the candidate has no strong-country signal nearby.
        rule_hint = (row[5] or "").strip() if row[5] else ""
        for ex_kw in split_keywords(weak_kw):
            if not ex_kw:
                continue
            entry = weak_agg.setdefault(ex_kw, {
                "weak_signal": ex_kw,
                "naive_country": "",
                "possible_countries": set(),
                "required_strong": set(),
                "guidance": "",
            })
            entry["possible_countries"].add(country_name)
            if rule_hint and rule_hint not in entry["guidance"]:
                entry["guidance"] = (entry["guidance"] + " | " + rule_hint).strip(" |")

    if dry_run:
        print(f"  [dry-run] would add {len(new_locations)} locations, "
              f"{len(new_aliases)} aliases, {len(new_landmarks)} landmarks, "
              f"{len(weak_agg)} weak signals (from country exclusion cols)")
        return {
            "locations": new_locations,
            "aliases": new_aliases,
            "landmarks": new_landmarks,
            "weak_agg": weak_agg,
        }

    # write out
    loc_fields, _ = read_rows(DATA_DIR / "locations.csv")
    al_fields, _ = read_rows(DATA_DIR / "aliases.csv")
    lm_fields, _ = read_rows(DATA_DIR / "landmark_index.csv")

    append_csv(DATA_DIR / "locations.csv", loc_fields, new_locations)
    append_csv(DATA_DIR / "aliases.csv", al_fields, new_aliases)
    append_csv(DATA_DIR / "landmark_index.csv", lm_fields, new_landmarks)

    return {
        "locations": new_locations,
        "aliases": new_aliases,
        "landmarks": new_landmarks,
        "weak_agg": weak_agg,
    }


def import_weak_signals(xlsx_path: Path, country_weak_agg: dict, dry_run: bool) -> int:
    """Merge two sources into weak_signals.csv:

    1. `弱信號排除詞庫` sheet — curated entries with naive_country and required_strong
    2. country_weak_agg — exclusion keywords from per-country rows, multiple countries possible

    Curated entries take precedence (they have richer fields).
    """
    wb = load_workbook(str(xlsx_path), data_only=True)
    ws = wb["弱信號排除詞庫"]
    out_path = DATA_DIR / "weak_signals.csv"

    merged: dict[str, dict] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        sig = (row[0] or "").strip()
        merged[sig] = {
            "weak_signal": sig,
            "naive_country": (row[1] or "").strip(),
            "possible_countries": (row[2] or "").strip(),
            "required_strong": (row[3] or "").strip(),
            "guidance": (row[4] or "").strip(),
            "source": "curated",
        }

    # P2-8 + P2-9: split bundled signals like "威尼斯、貢多拉、水都" into separate entries,
    # each with its own dict (deep copy avoids shared-state pollution when we later mutate
    # possible_countries). Remove the original bundled key to prevent duplicate rows
    # in weak_signals.csv.
    bundled_keys: list[str] = []
    expanded: dict[str, dict] = {}
    for sig, entry in list(merged.items()):
        parts = split_keywords(sig)
        if len(parts) <= 1:
            continue
        bundled_keys.append(sig)
        for part in parts:
            if part in merged or part in expanded:
                continue  # don't overwrite an already-curated single-word entry
            new_entry = dict(entry)
            new_entry["weak_signal"] = part
            expanded[part] = new_entry
    for k in bundled_keys:
        del merged[k]
    merged.update(expanded)

    for sig, agg in country_weak_agg.items():
        if sig in merged:
            # already curated; just augment possible_countries
            existing = merged[sig]
            existing_set = {c.strip() for c in existing.get("possible_countries", "").split("、") if c.strip()}
            merged_set = existing_set | agg["possible_countries"]
            existing["possible_countries"] = "、".join(sorted(merged_set))
            continue
        merged[sig] = {
            "weak_signal": sig,
            "naive_country": "",
            "possible_countries": "、".join(sorted(agg["possible_countries"])),
            "required_strong": "",
            "guidance": agg.get("guidance", "") or "需搭配該國強信號（城市/島/景點/品牌）",
            "source": "country_row",
        }

    # ensure every entry has the full schema fields (defaults to empty string)
    for entry in merged.values():
        entry.setdefault("min_context", "")
        entry.setdefault("source", "curated")
        # P1-6: ≤2-char weak signals (企鵝/夜市/夜景/古城/北角/極光...) are easily over-triggered
        # on unrelated docs. Auto-build a min_context: require at least one
        # possible_country name to appear in text. This dramatically cuts false positives
        # while preserving warnings on real DMs that name the country alongside the weak word.
        if len(entry["weak_signal"]) <= 2 and not entry["min_context"]:
            countries = [c.strip() for c in
                         (entry.get("possible_countries") or "").split("、") if c.strip()]
            if countries:
                # the keyword itself must also be present (already enforced by scan); we just
                # add: AT LEAST ONE country alias must co-occur in the document.
                entry["min_context"] = "|".join(countries)

    rows_out = sorted(merged.values(), key=lambda r: r["weak_signal"])

    if dry_run:
        print(f"  [dry-run] would write {len(rows_out)} weak signals to {out_path}")
        return len(rows_out)

    fieldnames = EXPECTED_FIELDS["weak_signals.csv"]
    with out_path.open("w", encoding="utf-8", newline="") as fh:
        writer = csv.DictWriter(fh, fieldnames=fieldnames, quoting=csv.QUOTE_MINIMAL)
        writer.writeheader()
        writer.writerows(rows_out)
    return len(rows_out)


def main() -> None:
    parser = argparse.ArgumentParser(description="Import Excel master into 5 CSVs")
    parser.add_argument("--xlsx", default=str(DEFAULT_XLSX), help="path to xlsx")
    parser.add_argument("--dry-run", action="store_true", help="report without writing")
    args = parser.parse_args()

    xlsx_path = Path(args.xlsx)
    if not xlsx_path.exists():
        raise SystemExit(f"xlsx not found: {xlsx_path}")

    kb = load_kb()
    print(f"Importing {xlsx_path}")
    print(f"Current KB: {len(kb.locations)} locations, {len(kb.aliases)} aliases, "
          f"{len(kb.rules)} rules, {len(kb.landmarks)} landmarks")

    result = import_country_sheet(xlsx_path, kb, args.dry_run)
    print(f"\nCountry sheet:")
    print(f"  + {len(result['locations'])} new countries")
    print(f"  + {len(result['aliases'])} new aliases")
    print(f"  + {len(result['landmarks'])} new landmarks")
    print(f"  + {len(result['weak_agg'])} weak signals from country exclusion cols")

    n_weak = import_weak_signals(xlsx_path, result["weak_agg"], args.dry_run)
    print(f"\nWeak signals (merged): {n_weak} entries -> weak_signals.csv")

    if not args.dry_run:
        print("\nDone. Run `python scripts/extract.py test` to verify regressions.")
        print("Run `python scripts/update_kb.py sync --to "
              "\"C:/Users/User/Desktop/2d3d/旅遊表/data\"` to mirror to working dir.")


if __name__ == "__main__":
    main()
